import io
import csv
import logging
from datetime import datetime

from aiogram import Router, F
from aiogram.types import (
    Message, CallbackQuery,
    InlineKeyboardMarkup, InlineKeyboardButton,
    BufferedInputFile
)
from keyboards.inline import main_menu, back_to_menu, admin_back_kb
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup

from database.db import (
    get_all_swaps, get_swaps_by_period, get_stats, get_recent_logs,
    get_top_pairs, get_average_amount, get_volume_by_period,
    get_all_currencies_admin, add_currency, toggle_currency,
    update_currency_min, delete_currency,
    get_all_users, get_user_swap_history,
    is_user_blocked, block_user, unblock_user,
    update_user_swaps_count  # Проверь, что добавил эту функцию в db.py
)
from config import ADMIN_ID

logger = logging.getLogger(__name__)
router = Router()


# ── FSM states ────────────────────────────────────────────────────────────────

class AdminStates(StatesGroup):
    # Currency add
    waiting_add_ticker   = State()
    waiting_add_network  = State()
    waiting_add_label    = State()
    waiting_add_min      = State()
    waiting_add_fiat     = State()
    # Currency edit
    waiting_edit_min     = State()
    # Custom date export
    waiting_date_from    = State()
    waiting_date_to      = State()
    # User lookup
    waiting_user_id      = State()
    # Block
    waiting_block_id     = State()
    waiting_block_reason = State()
    waiting_unblock_id   = State()


def is_admin(user_id: int) -> bool:
    return user_id == ADMIN_ID


# ── Keyboards ─────────────────────────────────────────────────────────────────

def admin_main_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="💰 Currencies",    callback_data="adm_currencies")],
        [InlineKeyboardButton(text="📋 Swaps",         callback_data="adm_swaps_menu")],
        [InlineKeyboardButton(text="👥 Users",         callback_data="adm_users_menu")],
        [InlineKeyboardButton(text="📈 Top pairs",     callback_data="adm_pairs")],
        [InlineKeyboardButton(text="📋 Logs",          callback_data="adm_logs")],
    ])


def swaps_period_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="📅 Today",    callback_data="adm_swaps_period_1"),
            InlineKeyboardButton(text="📅 7 days",   callback_data="adm_swaps_period_7"),
        ],
        [
            InlineKeyboardButton(text="📅 30 days",  callback_data="adm_swaps_period_30"),
            InlineKeyboardButton(text="📅 All time", callback_data="adm_swaps_period_0"),
        ],
        [InlineKeyboardButton(text="📆 Custom dates", callback_data="adm_swaps_custom")],
        [InlineKeyboardButton(text="⬅️ Back",          callback_data="adm_back")],
    ])


def export_keyboard(period_tag: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="⬇️ Export CSV",   callback_data=f"adm_export_csv_{period_tag}"),
            InlineKeyboardButton(text="⬇️ Export Excel", callback_data=f"adm_export_xlsx_{period_tag}"),
        ],
        [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_swaps_menu")],
    ])


def users_menu_keyboard() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [InlineKeyboardButton(text="👥 All users",       callback_data="adm_users_list")],
        [InlineKeyboardButton(text="🔍 User by ID",      callback_data="adm_user_lookup")],
        [InlineKeyboardButton(text="🚫 Block user",      callback_data="adm_block_start")],
        [InlineKeyboardButton(text="✅ Unblock user",    callback_data="adm_unblock_start")],
        [InlineKeyboardButton(text="⬅️ Back",            callback_data="adm_back")],
    ])

# Вспомогательная функция для меню управления рангами
def get_user_manage_kb(target_id: int) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[
        [
            InlineKeyboardButton(text="🥈 Silver (10)", callback_data=f"set_swaps_{target_id}_10"),
            InlineKeyboardButton(text="🥇 Gold (25)", callback_data=f"set_swaps_{target_id}_25")
        ],
        [
            InlineKeyboardButton(text="💎 Diamond (50)", callback_data=f"set_swaps_{target_id}_50"),
            InlineKeyboardButton(text="👑 VIP (100)", callback_data=f"set_swaps_{target_id}_100")
        ],
        [
            InlineKeyboardButton(text="🚫 Block", callback_data=f"adm_block_confirm_{target_id}"),
            InlineKeyboardButton(text="✅ Unblock", callback_data=f"adm_unblock_confirm_{target_id}")
        ],
        [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_users_menu")]
    ])


# ── /admin ────────────────────────────────────────────────────────────────────

@router.message(Command("admin"))
async def cmd_admin(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    stats = await get_stats()
    avg   = await get_average_amount()
    await message.answer(
        f"👑 <b>Admin Panel</b>\n\n"
        f"📊 Total: <b>{stats['total']}</b>  |  Users: <b>{stats['users']}</b>  |  Avg: <b>{avg}</b>\n\n"
        f"📅 Today: <b>{stats['today']}</b>  |  Week: <b>{stats['week']}</b>  |  Month: <b>{stats['month']}</b>\n\n"
        f"✅ Finished: <b>{stats['finished']}</b>  "
        f"⏳ Pending: <b>{stats['pending']}</b>  "
        f"❌ Failed: <b>{stats['failed']}</b>",
        reply_markup=admin_main_keyboard()
    )

@router.callback_query(F.data == "admin_stats")
async def admin_stats(callback: CallbackQuery):
    from database.db import get_stats
    s = await get_stats()
    
    text = (
        "📊 <b>Global Stats</b>\n\n"
        f"👥 Users: {s['registrations_total']}\n"
        f"🆕 Regs (Today/Week): {s['reg_today']} / {s['reg_week']}\n"
        f"🔄 Total Swaps: {s['total']}\n"
        f"✅ Finished: {s['finished']}\n"
        f"⏳ Pending: {s['pending']}\n"
        f"❌ Failed: {s['failed']}\n"
    )
    await callback.message.edit_text(text, reply_markup=admin_back_kb())
    
# 1. Вызываем меню управления юзером (через команду)
@router.message(F.text.startswith("/user_manage"))
async def admin_manage_user(message: Message):
    if not is_admin(message.from_user.id): return
    try:
        parts = message.text.split()
        if len(parts) < 2:
            return await message.answer("Usage: <code>/user_manage [USER_ID]</code>")
        target_id = int(parts[1])
        await message.answer(f"⚙️ <b>Managing User:</b> <code>{target_id}</code>", reply_markup=get_user_manage_kb(target_id))
    except ValueError:
        await message.answer("❌ Invalid User ID.")

# 2. Обработка кнопки из канала
@router.callback_query(F.data.startswith("admin_manage_"))
async def admin_manage_callback(callback: CallbackQuery):
    if not is_admin(callback.from_user.id): return
    target_id = int(callback.data.split("_")[2])
    await callback.message.answer(f"⚙️ <b>Managing User:</b> <code>{target_id}</code>", reply_markup=get_user_manage_kb(target_id))
    await callback.answer()

# 3. Установка ранга (сделок)
@router.callback_query(F.data.startswith("set_swaps_"))
async def admin_set_swaps(callback: CallbackQuery):
    if not is_admin(callback.from_user.id): return
    parts = callback.data.split("_")
    target_id = int(parts[2])
    count = int(parts[3])
    
    success = await update_user_swaps_count(target_id, count)
    if success:
        await callback.answer(f"✅ User rank updated!", show_alert=True)
        await callback.message.edit_text(f"✅ User <code>{target_id}</code> now has <b>{count}</b> swaps.")
    else:
        await callback.answer("❌ DB Error", show_alert=True)


@router.callback_query(F.data == "adm_back")
async def cb_adm_back(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    stats = await get_stats()
    avg   = await get_average_amount()
    await callback.message.edit_text(
        f"👑 <b>Admin Panel</b>\n\n"
        f"📊 Total: <b>{stats['total']}</b>  |  Users: <b>{stats['users']}</b>  |  Avg: <b>{avg}</b>\n\n"
        f"📅 Today: <b>{stats['today']}</b>  |  Week: <b>{stats['week']}</b>  |  Month: <b>{stats['month']}</b>\n\n"
        f"✅ Finished: <b>{stats['finished']}</b>  "
        f"⏳ Pending: <b>{stats['pending']}</b>  "
        f"❌ Failed: <b>{stats['failed']}</b>",
        reply_markup=admin_main_keyboard()
    )


# ── Swaps with period filter ───────────────────────────────────────────────────

@router.message(Command("admin_swaps"))
async def cmd_admin_swaps(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    await message.answer("📋 <b>Swaps — choose period:</b>", reply_markup=swaps_period_keyboard())


@router.callback_query(F.data == "adm_swaps_menu")
async def cb_swaps_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await callback.message.edit_text(
        "📋 <b>Swaps — choose period:</b>",
        reply_markup=swaps_period_keyboard()
    )


@router.callback_query(F.data.startswith("adm_swaps_period_"))
async def cb_swaps_period(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    days_str = callback.data.split("_")[-1]
    days     = int(days_str)

    if days == 0:
        swaps    = await get_swaps_by_period()
        label    = "All time"
        tag      = "all"
    elif days == 1:
        swaps    = await get_swaps_by_period(days=1)
        label    = "Today"
        tag      = "1d"
    else:
        swaps    = await get_swaps_by_period(days=days)
        label    = f"Last {days} days"
        tag      = f"{days}d"

    if not swaps:
        await callback.message.edit_text(
            f"📭 No swaps for: <b>{label}</b>",
            reply_markup=swaps_period_keyboard()
        )
        return

    text = f"📋 <b>Swaps — {label}</b> ({len(swaps)} total)\n\n"
    for s in swaps[:15]:
        text += (
            f"👤 <code>{s['user_id']}</code> | "
            f"{s['amount_from']} {s['currency_from'].upper()} → "
            f"{s['currency_to'].upper()} | {s['status']}\n"
            f"   <code>{s['exchange_id']}</code>\n\n"
        )
    if len(swaps) > 15:
        text += f"<i>...and {len(swaps) - 15} more. Export for full list.</i>\n"

    await callback.message.edit_text(text, reply_markup=export_keyboard(tag))


# Custom date range
@router.callback_query(F.data == "adm_swaps_custom")
async def cb_swaps_custom(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await state.set_state(AdminStates.waiting_date_from)
    await callback.message.edit_text(
        "📆 <b>Custom date range</b>\n\n"
        "Enter <b>start date</b> in format <code>YYYY-MM-DD</code>\n"
        "Example: <code>2025-01-01</code>\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_date_from)
async def process_date_from(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        datetime.strptime(message.text.strip(), "%Y-%m-%d")
    except ValueError:
        await message.answer("⚠️ Invalid format. Use <code>YYYY-MM-DD</code>. Try again or /cancel")
        return
    await state.update_data(date_from=message.text.strip())
    await state.set_state(AdminStates.waiting_date_to)
    await message.answer(
        "📆 Now enter <b>end date</b> in format <code>YYYY-MM-DD</code>\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_date_to)
async def process_date_to(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        datetime.strptime(message.text.strip(), "%Y-%m-%d")
    except ValueError:
        await message.answer("⚠️ Invalid format. Use <code>YYYY-MM-DD</code>. Try again or /cancel")
        return
    data      = await state.get_data()
    date_from = data["date_from"]
    date_to   = message.text.strip()
    await state.update_data(date_to=date_to)
    await state.clear()

    swaps = await get_swaps_by_period(date_from=date_from, date_to=date_to)
    tag   = f"custom_{date_from}_{date_to}"
    label = f"{date_from} → {date_to}"

    if not swaps:
        await message.answer(f"📭 No swaps for: <b>{label}</b>")
        return

    text = f"📋 <b>Swaps — {label}</b> ({len(swaps)} total)\n\n"
    for s in swaps[:15]:
        text += (
            f"👤 <code>{s['user_id']}</code> | "
            f"{s['amount_from']} {s['currency_from'].upper()} → "
            f"{s['currency_to'].upper()} | {s['status']}\n"
            f"   <code>{s['exchange_id']}</code>\n\n"
        )
    if len(swaps) > 15:
        text += f"<i>...and {len(swaps) - 15} more. Export for full list.</i>\n"

    await message.answer(text, reply_markup=export_keyboard(tag))


# ── Export CSV / Excel ─────────────────────────────────────────────────────────

async def _get_swaps_for_tag(tag: str) -> tuple[list, str]:
    """Parse period tag and return (swaps, label)."""
    if tag == "all":
        return await get_swaps_by_period(), "all_time"
    if tag == "1d":
        return await get_swaps_by_period(days=1), "today"
    if tag == "7d":
        return await get_swaps_by_period(days=7), "7_days"
    if tag == "30d":
        return await get_swaps_by_period(days=30), "30_days"
    if tag.startswith("custom_"):
        parts     = tag.split("_", 1)[1]
        date_from, date_to = parts.split("_")
        return await get_swaps_by_period(date_from=date_from, date_to=date_to), f"{date_from}_{date_to}"
    return await get_swaps_by_period(), "export"


@router.callback_query(F.data.startswith("adm_export_csv_"))
async def cb_export_csv(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer("Generating CSV...")
    tag    = callback.data[len("adm_export_csv_"):]
    swaps, label = await _get_swaps_for_tag(tag)

    if not swaps:
        await callback.message.answer("📭 No data to export.")
        return

    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=[
        "id", "user_id", "exchange_id",
        "currency_from", "currency_to",
        "amount_from", "amount_to",
        "address_to", "status", "created_at"
    ])
    writer.writeheader()
    writer.writerows(swaps)
    buf.seek(0)

    filename = f"swaps_{label}_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    await callback.message.answer_document(
        document=BufferedInputFile(buf.getvalue().encode("utf-8"), filename=filename),
        caption=f"📊 Swaps export — {label.replace('_', ' ')} ({len(swaps)} rows)"
    )


@router.callback_query(F.data.startswith("adm_export_xlsx_"))
async def cb_export_xlsx(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer("Generating Excel...")
    tag    = callback.data[len("adm_export_xlsx_"):]
    swaps, label = await _get_swaps_for_tag(tag)

    if not swaps:
        await callback.message.answer("📭 No data to export.")
        return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        await callback.message.answer("⚠️ openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Swaps"

    headers = ["ID", "User ID", "Exchange ID", "From", "To",
               "Amount From", "Amount To", "Address To", "Status", "Created At"]
    header_fill = PatternFill("solid", fgColor="1e1e2e")
    header_font = Font(bold=True, color="7c6af7")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")

    fields = ["id", "user_id", "exchange_id", "currency_from", "currency_to",
              "amount_from", "amount_to", "address_to", "status", "created_at"]
    for row_idx, swap in enumerate(swaps, 2):
        for col_idx, field in enumerate(fields, 1):
            ws.cell(row=row_idx, column=col_idx, value=swap.get(field, ""))

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"swaps_{label}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await callback.message.answer_document(
        document=BufferedInputFile(buf.getvalue(), filename=filename),
        caption=f"📊 Swaps export — {label.replace('_', ' ')} ({len(swaps)} rows)"
    )


# ── User management ────────────────────────────────────────────────────────────

@router.message(Command("admin_users"))
async def cmd_admin_users(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    await message.answer("👥 <b>User management:</b>", reply_markup=users_menu_keyboard())


@router.callback_query(F.data == "adm_users_menu")
async def cb_users_menu(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await callback.message.edit_text(
        "👥 <b>User management:</b>",
        reply_markup=users_menu_keyboard()
    )


@router.callback_query(F.data == "adm_users_list")
async def cb_users_list(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    users = await get_all_users()
    if not users:
        await callback.message.edit_text(
            "📭 No users yet.",
            reply_markup=users_menu_keyboard()
        )
        return

    text = f"👥 <b>Users ({len(users)} total)</b>\n\n"
    for u in users[:20]:
        blocked = " 🚫" if u["is_blocked"] else ""
        text += (
            f"<code>{u['user_id']}</code>{blocked}\n"
            f"  Swaps: <b>{u['swap_count']}</b>  |  "
            f"Vol: <b>{round(u['total_volume'] or 0, 4)}</b>  |  "
            f"Last: <b>{str(u['last_swap'])[:10]}</b>\n\n"
        )
    if len(users) > 20:
        text += f"<i>...and {len(users) - 20} more</i>"

    await callback.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="🔍 Lookup user", callback_data="adm_user_lookup")],
            [InlineKeyboardButton(text="⬅️ Back",        callback_data="adm_users_menu")],
        ])
    )


# User lookup by ID
@router.callback_query(F.data == "adm_user_lookup")
async def cb_user_lookup(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await state.set_state(AdminStates.waiting_user_id)
    await callback.message.edit_text(
        "🔍 <b>User lookup</b>\n\n"
        "Enter Telegram user ID:\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_user_id)
async def process_user_lookup(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer("⚠️ Invalid ID. Enter a numeric Telegram user ID or /cancel")
        return
    await state.clear()

    swaps   = await get_user_swap_history(user_id, limit=20)
    blocked = await is_user_blocked(user_id)
    status  = "🚫 Blocked" if blocked else "✅ Active"

    if not swaps:
        await message.answer(
            f"👤 User <code>{user_id}</code>\n"
            f"Status: {status}\n\n"
            f"📭 No swaps found.",
            reply_markup=_user_action_keyboard(user_id, blocked)
        )
        return

    text = (
        f"👤 User <code>{user_id}</code>  |  {status}\n"
        f"Total swaps: <b>{len(swaps)}</b>\n\n"
    )
    for s in swaps[:10]:
        text += (
            f"• {s['amount_from']} {s['currency_from'].upper()} → "
            f"{s['currency_to'].upper()}  [{s['status']}]\n"
            f"  <code>{s['exchange_id']}</code>  {str(s['created_at'])[:10]}\n\n"
        )
    if len(swaps) > 10:
        text += f"<i>...and {len(swaps) - 10} more</i>"

    await message.answer(text, reply_markup=_user_action_keyboard(user_id, blocked))


def _user_action_keyboard(user_id: int, is_blocked: bool) -> InlineKeyboardMarkup:
    if is_blocked:
        toggle_btn = InlineKeyboardButton(
            text="✅ Unblock", callback_data=f"adm_unblock_confirm_{user_id}"
        )
    else:
        toggle_btn = InlineKeyboardButton(
            text="🚫 Block", callback_data=f"adm_block_confirm_{user_id}"
        )
    return InlineKeyboardMarkup(inline_keyboard=[
        [toggle_btn],
        [InlineKeyboardButton(text="⚙️ Manage Rank", callback_data=f"admin_manage_{user_id}")], # Добавил быстрый переход к рангам
        [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_users_menu")],
    ])


# Block
@router.callback_query(F.data == "adm_block_start")
async def cb_block_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await state.set_state(AdminStates.waiting_block_id)
    await callback.message.edit_text(
        "🚫 <b>Block user</b>\n\nEnter Telegram user ID:\n\n<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_block_id)
async def process_block_id(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer("⚠️ Invalid ID. Try again or /cancel")
        return
    await state.update_data(block_user_id=user_id)
    await state.set_state(AdminStates.waiting_block_reason)
    await message.answer(
        f"🚫 Block user <code>{user_id}</code>\n\n"
        "Enter reason (or send <code>-</code> to skip):\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_block_reason)
async def process_block_reason(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    reason  = "" if message.text.strip() == "-" else message.text.strip()
    data    = await state.get_data()
    user_id = data["block_user_id"]
    await state.clear()

    ok = await block_user(user_id, reason)
    if ok:
        await message.answer(f"🚫 User <code>{user_id}</code> blocked.")
    else:
        await message.answer(f"⚠️ User <code>{user_id}</code> is already blocked.")


@router.callback_query(F.data.startswith("adm_block_confirm_"))
async def cb_block_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    user_id = int(callback.data.split("_")[-1])
    ok      = await block_user(user_id)
    await callback.answer("🚫 Blocked" if ok else "Already blocked", show_alert=True)
    swaps   = await get_user_swap_history(user_id, limit=1)
    blocked = await is_user_blocked(user_id)
    await callback.message.edit_reply_markup(
        reply_markup=_user_action_keyboard(user_id, blocked)
    )


# Unblock
@router.callback_query(F.data == "adm_unblock_start")
async def cb_unblock_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await state.set_state(AdminStates.waiting_unblock_id)
    await callback.message.edit_text(
        "✅ <b>Unblock user</b>\n\nEnter Telegram user ID:\n\n<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_unblock_id)
async def process_unblock_id(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        user_id = int(message.text.strip())
    except ValueError:
        await message.answer("⚠️ Invalid ID. Try again or /cancel")
        return
    await state.clear()
    ok = await unblock_user(user_id)
    if ok:
        await message.answer(f"✅ User <code>{user_id}</code> unblocked.")
    else:
        await message.answer(f"⚠️ User <code>{user_id}</code> was not blocked.")


@router.callback_query(F.data.startswith("adm_unblock_confirm_"))
async def cb_unblock_confirm(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    user_id = int(callback.data.split("_")[-1])
    ok      = await unblock_user(user_id)
    await callback.answer("✅ Unblocked" if ok else "Not blocked", show_alert=True)
    blocked = await is_user_blocked(user_id)
    await callback.message.edit_reply_markup(
        reply_markup=_user_action_keyboard(user_id, blocked)
    )


# ── Currency management ────────────────────────────────────────────────────────

async def _currencies_text_and_kb() -> tuple[str, InlineKeyboardMarkup]:
    currencies = await get_all_currencies_admin()
    if not currencies:
        text = "💰 <b>Currencies</b>\n\nNo currencies configured yet."
        kb   = InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="➕ Add", callback_data="adm_add_currency")],
            [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_back")],
        ])
        return text, kb

    lines = []
    for c in currencies:
        st    = "✅" if c["is_active"] else "❌"
        ftype = "FIAT" if c["is_fiat"] else "CRYPTO"
        lines.append(
            f"{st} <b>{c['label']}</b> — "
            f"<code>{c['ticker']}/{c['network']}</code> "
            f"[{ftype}] min: <b>{c['min_amount']}</b>"
        )
    text = "💰 <b>Currencies</b>\n\n" + "\n".join(lines)

    rows = []
    for c in currencies:
        toggle_txt = "🔴 Disable" if c["is_active"] else "🟢 Enable"
        rows.append([
            InlineKeyboardButton(text=c["label"],   callback_data=f"adm_cur_info_{c['id']}"),
            InlineKeyboardButton(text=toggle_txt,   callback_data=f"adm_cur_toggle_{c['id']}"),
            InlineKeyboardButton(text="✏️ Min",     callback_data=f"adm_cur_editmin_{c['id']}"),
            InlineKeyboardButton(text="🗑",         callback_data=f"adm_cur_delete_{c['id']}"),
        ])
    rows.append([InlineKeyboardButton(text="➕ Add currency", callback_data="adm_add_currency")])
    rows.append([InlineKeyboardButton(text="⬅️ Back",         callback_data="adm_back")])
    return text, InlineKeyboardMarkup(inline_keyboard=rows)


@router.message(Command("admin_currencies"))
async def cmd_admin_currencies(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    text, kb = await _currencies_text_and_kb()
    await message.answer(text, reply_markup=kb)


@router.callback_query(F.data == "adm_currencies")
async def cb_currencies(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    text, kb = await _currencies_text_and_kb()
    await callback.message.edit_text(text, reply_markup=kb)


@router.callback_query(F.data.startswith("adm_cur_toggle_"))
async def cb_toggle(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    cur_id    = int(callback.data.split("_")[-1])
    new_state = await toggle_currency(cur_id)
    await callback.answer("✅ Enabled" if new_state else "❌ Disabled", show_alert=True)
    text, kb  = await _currencies_text_and_kb()
    await callback.message.edit_text(text, reply_markup=kb)


@router.callback_query(F.data.startswith("adm_cur_editmin_"))
async def cb_editmin(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    cur_id = int(callback.data.split("_")[-1])
    await state.set_state(AdminStates.waiting_edit_min)
    await state.update_data(edit_currency_id=cur_id)
    await callback.message.edit_text(
        "✏️ <b>Edit minimum amount</b>\n\n"
        "Enter new value, e.g. <code>0.001</code>\n\n<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_edit_min)
async def process_edit_min(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        new_min = float(message.text.replace(",", "."))
        if new_min < 0:
            raise ValueError
    except ValueError:
        await message.answer("⚠️ Invalid amount. Try again or /cancel")
        return
    data = await state.get_data()
    await update_currency_min(data["edit_currency_id"], new_min)
    await state.clear()
    await message.answer(f"✅ Minimum updated to <b>{new_min}</b>")
    text, kb = await _currencies_text_and_kb()
    await message.answer(text, reply_markup=kb)


@router.callback_query(F.data.startswith("adm_cur_delete_"))
async def cb_delete(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    cur_id = int(callback.data.split("_")[-1])
    await callback.message.edit_text(
        "🗑 <b>Confirm deletion?</b>\n\nThis cannot be undone.",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="✅ Yes, delete", callback_data=f"adm_cur_confirmdelete_{cur_id}"),
                InlineKeyboardButton(text="❌ Cancel",       callback_data="adm_currencies"),
            ]
        ])
    )


@router.callback_query(F.data.startswith("adm_cur_confirmdelete_"))
async def cb_confirm_delete(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    cur_id = int(callback.data.split("_")[-1])
    await delete_currency(cur_id)
    await callback.answer("🗑 Deleted", show_alert=True)
    text, kb = await _currencies_text_and_kb()
    await callback.message.edit_text(text, reply_markup=kb)


@router.callback_query(F.data == "adm_add_currency")
async def cb_add_start(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    await state.set_state(AdminStates.waiting_add_ticker)
    await callback.message.edit_text(
        "➕ <b>Add currency — Step 1/5</b>\n\n"
        "Enter the <b>ticker</b>\n"
        "Example: <code>btc</code>, <code>eth</code>, <code>matic</code>\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_add_ticker)
async def process_add_ticker(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.update_data(new_ticker=message.text.strip().lower())
    await state.set_state(AdminStates.waiting_add_network)
    await message.answer(
        "➕ <b>Step 2/5</b>\n\nEnter the <b>network</b>\n"
        "Example: <code>eth</code>, <code>bsc</code>, <code>trx</code>\n"
        "For fiat use same as ticker: <code>usd</code>, <code>eur</code>\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_add_network)
async def process_add_network(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.update_data(new_network=message.text.strip().lower())
    await state.set_state(AdminStates.waiting_add_label)
    await message.answer(
        "➕ <b>Step 3/5</b>\n\nEnter the <b>display label</b>\n"
        "Example: <code>MATIC</code>, <code>USDT (TRC20)</code>\n\n"
        "<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_add_label)
async def process_add_label(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    await state.update_data(new_label=message.text.strip())
    await state.set_state(AdminStates.waiting_add_min)
    await message.answer(
        "➕ <b>Step 4/5</b>\n\nEnter the <b>minimum amount</b>\n"
        "Example: <code>5.0</code>\n\n<i>/cancel to abort</i>"
    )


@router.message(AdminStates.waiting_add_min)
async def process_add_min(message: Message, state: FSMContext):
    if not is_admin(message.from_user.id):
        return
    try:
        min_amount = float(message.text.replace(",", "."))
        if min_amount < 0:
            raise ValueError
    except ValueError:
        await message.answer("⚠️ Invalid amount. Try again or /cancel")
        return
    await state.update_data(new_min=min_amount)
    await state.set_state(AdminStates.waiting_add_fiat)
    await message.answer(
        "➕ <b>Step 5/5</b>\n\nIs this a <b>fiat</b> currency?",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [
                InlineKeyboardButton(text="💳 Yes — fiat",  callback_data="adm_add_fiat_yes"),
                InlineKeyboardButton(text="🔗 No — crypto", callback_data="adm_add_fiat_no"),
            ]
        ])
    )


@router.callback_query(F.data.startswith("adm_add_fiat_"))
async def process_add_fiat(callback: CallbackQuery, state: FSMContext):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    is_fiat = callback.data == "adm_add_fiat_yes"
    data    = await state.get_data()
    await state.clear()
    cur_id  = await add_currency(
        ticker=data["new_ticker"], network=data["new_network"],
        label=data["new_label"],   min_amount=data["new_min"],
        is_fiat=is_fiat,
    )
    ftype = "fiat" if is_fiat else "crypto"
    await callback.message.edit_text(
        f"✅ <b>Currency added!</b>\n\n"
        f"Label: <b>{data['new_label']}</b>\n"
        f"Ticker/Network: <code>{data['new_ticker']}/{data['new_network']}</code>\n"
        f"Min: <b>{data['new_min']}</b> | Type: <b>{ftype}</b> | ID: <code>{cur_id}</code>"
    )
    text, kb = await _currencies_text_and_kb()
    await callback.message.answer(text, reply_markup=kb)


# ── Pairs ──────────────────────────────────────────────────────────────────────

@router.message(Command("admin_pairs"))
async def cmd_admin_pairs(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    pairs = await get_top_pairs()
    if not pairs:
        await message.answer("📭 No finished exchanges yet.")
        return
    medals = ["🥇", "🥈", "🥉", "4.", "5."]
    text   = "📈 <b>Top pairs:</b>\n\n"
    for i, p in enumerate(pairs):
        m = medals[i] if i < len(medals) else f"{i+1}."
        text += (
            f"{m} <b>{p['currency_from'].upper()} → {p['currency_to'].upper()}</b>\n"
            f"   Exchanges: {p['count']} | Volume: {round(p['volume'], 4)}\n\n"
        )
    await message.answer(text)


@router.callback_query(F.data == "adm_pairs")
async def cb_adm_pairs(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    pairs = await get_top_pairs()
    if not pairs:
        await callback.message.edit_text(
            "📭 No finished exchanges yet.",
            reply_markup=InlineKeyboardMarkup(inline_keyboard=[
                [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_back")]
            ])
        )
        return
    medals = ["🥇", "🥈", "🥉", "4.", "5."]
    text   = "📈 <b>Top pairs:</b>\n\n"
    for i, p in enumerate(pairs):
        m = medals[i] if i < len(medals) else f"{i+1}."
        text += (
            f"{m} <b>{p['currency_from'].upper()} → {p['currency_to'].upper()}</b>\n"
            f"   Exchanges: {p['count']} | Volume: {round(p['volume'], 4)}\n\n"
        )
    await callback.message.edit_text(
        text,
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_back")]
        ])
    )


# ── Logs ───────────────────────────────────────────────────────────────────────

@router.message(Command("admin_logs"))
async def cmd_admin_logs(message: Message):
    if not is_admin(message.from_user.id):
        await message.answer("⛔ No access.")
        return
    lines = await get_recent_logs(20)
    if not lines:
        await message.answer("📭 No logs yet.")
        return
    text = "\n".join(lines)
    for chunk in [text[i:i+3500] for i in range(0, len(text), 3500)]:
        await message.answer(f"📋 <b>Recent logs:</b>\n\n<pre>{chunk}</pre>")


@router.callback_query(F.data == "adm_logs")
async def cb_adm_logs(callback: CallbackQuery):
    if not is_admin(callback.from_user.id):
        return
    await callback.answer()
    lines = await get_recent_logs(20)
    text  = "\n".join(lines) if lines else "No logs yet."
    await callback.message.edit_text(
        f"<pre>{text[:3500]}</pre>",
        reply_markup=InlineKeyboardMarkup(inline_keyboard=[
            [InlineKeyboardButton(text="⬅️ Back", callback_data="adm_back")]
        ])
    )